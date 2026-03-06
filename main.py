"""
CAMTEL Budget Management System v12
- Full directions list (50+)
- Per-user direction assignments
- Amount formatting with spaces: 300 000 000
- Print fiches (2 per A4)
- Budget line filter by direction
- DISPONIBLE OUI/NON
- AVIS DE LA DCF
- Initiator name on transactions
- Excel export templates
- Admin user management
"""
import os, io, csv, json, hashlib, logging, re
from datetime import date, datetime
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from itsdangerous import URLSafeSerializer, BadSignature
from sqlalchemy import (create_engine, func, Column, Integer, String,
                        Float, Boolean, DateTime, Text, UniqueConstraint)
from sqlalchemy.orm import sessionmaker, declarative_base, Session
from sqlalchemy.sql import func as sqlfunc

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

log = logging.getLogger("camtel")
logging.basicConfig(level=logging.INFO)

# ── DATABASE ──────────────────────────────────────────────────────────
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./camtel.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
IS_SQLITE = DATABASE_URL.startswith("sqlite")

if IS_SQLITE:
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False}, pool_pre_ping=True)
else:
    engine = create_engine(DATABASE_URL, pool_pre_ping=True,
                           pool_size=5, max_overflow=10, pool_recycle=1800)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False)
Base = declarative_base()

def get_db():
    db = SessionLocal()
    try:    yield db
    finally: db.close()

# ── MODELS ────────────────────────────────────────────────────────────
class User(Base):
    __tablename__ = "users"
    id           = Column(Integer, primary_key=True)
    username     = Column(String(80), unique=True, nullable=False, index=True)
    password     = Column(String(128), nullable=False)
    full_name    = Column(String(200), default="")
    role         = Column(String(20), default="agent")
    email        = Column(String(200), default="")
    is_active    = Column(Boolean, default=True)
    directions   = Column(Text, default="[]")  # JSON list of assigned directions

class BudgetLine(Base):
    __tablename__ = "budget_lines"
    __table_args__ = (UniqueConstraint("year", "direction", "imputation"),)
    id         = Column(Integer, primary_key=True)
    year       = Column(Integer, nullable=False, index=True)
    direction  = Column(String(50), nullable=False, index=True)
    imputation = Column(String(50), nullable=False, index=True)
    libelle    = Column(String(500), default="")
    nature     = Column(String(100), default="DEPENSE COURANTE")
    budget_cp  = Column(Float, default=0.0)

class Transaction(Base):
    __tablename__ = "transactions"
    id               = Column(Integer, primary_key=True)
    date_reception   = Column(String(20), nullable=False, index=True)
    direction        = Column(String(50), default="", index=True)
    imputation       = Column(String(50), default="", index=True)
    intitule         = Column(String(500), default="")
    libelle          = Column(String(500), default="")
    objet            = Column(Text, default="")  # description/objet
    nature           = Column(String(100), default="DEPENSE COURANTE")
    designation      = Column(String(20), default="")
    code_ref         = Column(String(300), default="")
    montant          = Column(Float, default=0.0)
    year             = Column(Integer, nullable=False, index=True)
    status           = Column(String(20), default="validated")
    statut_budget    = Column(String(20), default="OK")
    created_by       = Column(String(80), default="")
    created_by_name  = Column(String(200), default="")
    created_at       = Column(DateTime, default=datetime.utcnow)

# ── CONSTANTS ────────────────────────────────────────────────────────
ALL_DIRS = [
    "BUM","BUT","BUF","DG","DRH","DICOM","DIRCAB","DCRA","DAMR","DC",
    "DNQ","DAS","DFA","DAJR","DAP","DR","DS","DSPI","DSIR","DOP","DT",
    "DCF","DCRM","DRLM","RRSM","RREM","RROM","RRNOM","RRSOM","RRAM",
    "RRNM","RRENM","DCRF","DRLF","RRSF","RREF","RROF","RRNOF","RRSOF",
    "RRAF","RRNF","RRENF","DCRT","DRLT","RRNOT","RRENT"
]
MONTHS_FR = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

# ── AUTH ──────────────────────────────────────────────────────────────
SECRET_KEY = os.getenv("SECRET_KEY", "camtel-dev-secret")
ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASS = os.getenv("ADMIN_PASS", "admin123")
SER = URLSafeSerializer(SECRET_KEY, salt="camtel-v12")

def _hash(p): return hashlib.sha256(p.encode()).hexdigest()

def current_user(request: Request):
    token = request.cookies.get("session", "")
    auth  = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        b = auth[7:].strip()
        if b: token = b
    if not token: raise HTTPException(401, "Not authenticated")
    try:    return SER.loads(token)
    except: raise HTTPException(401, "Invalid token")

# ── HELPERS ───────────────────────────────────────────────────────────
def fmt_amount(n) -> str:
    """Format number with space thousands separator: 300 000 000"""
    if n is None: return "0"
    return f"{int(n):,}".replace(",", " ")

def clean_amount(s) -> float:
    s = str(s or "0").strip()
    for ch in ("\xa0","\u202f","\u00a0"," ","\t"): s = s.replace(ch, "")
    if not s or s in ("-",""): return 0.0
    if s.count(",") > 1: s = s.replace(",", "")
    elif s.count(",") == 1:
        p = s.split(",")
        s = s.replace(",","") if len(p[1])==3 else s.replace(",",".")
    try: return float(s)
    except: return 0.0

def norm_date(s) -> str:
    s = str(s or "").strip()
    if "/" in s:
        p = s.split("/")
        if len(p)==3 and len(p[2])==4: return f"{p[2]}-{int(p[1]):02d}-{int(p[0]):02d}"
    return s if len(s)==10 and "-" in s else date.today().isoformat()

def decode_csv(raw: bytes) -> str:
    for enc in ("utf-8-sig","utf-8","latin-1","cp1252"):
        try: return raw.decode(enc)
        except: pass
    return raw.decode("latin-1", errors="replace")

def read_file_rows(raw: bytes, filename: str = "") -> list:
    fname = filename.lower()
    if (fname.endswith(".xlsx") or fname.endswith(".xls")) and HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c).strip() if c is not None else "" for c in row])
            wb.close()
            return rows
        except Exception as e:
            log.warning("xlsx read failed: %s", e)
    txt = decode_csv(raw)
    return list(csv.reader(txt.splitlines()))

def extract_imp(col6: str) -> str:
    if ">>>" in col6:
        part = col6.split(">>>")[0]
        for seg in reversed(part.split("/")):
            seg = seg.strip()
            if re.match(r"^\d{6,10}$", seg): return seg
    return ""

def find_header_row(rows):
    for i, row in enumerate(rows):
        joined = " ".join(str(c) for c in row).upper()
        if "DATE ENGAGEMENT" in joined or ("DIRECTION" in joined and "MONTANT" in joined):
            return i
    return 0

def parse_camtel_tx(row):
    if len(row) < 2: return None
    direction = str(row[1]).strip().upper()
    if not direction or direction in ("DIRECTION","TOTAL","SOUS-TOTAL","CE","","CODES","NONE"): return None
    if direction.startswith("DATE") or direction.startswith("CODE"): return None
    montant = clean_amount(row[7]) if len(row) > 7 else 0.0
    if montant <= 0: return None
    return {
        "date":        norm_date(str(row[0]).strip()),
        "direction":   direction,
        "intitule":    str(row[2]).strip() if len(row)>2 else "",
        "libelle":     str(row[3]).strip() if len(row)>3 else "",
        "nature":      str(row[4]).strip() if len(row)>4 and str(row[4]).strip() else "DEPENSE COURANTE",
        "designation": str(row[5]).strip() if len(row)>5 else "",
        "code_ref":    str(row[6]).strip() if len(row)>6 else "",
        "imputation":  extract_imp(str(row[6]).strip()) if len(row)>6 else "",
        "montant":     montant,
    }

def get_user_directions(db: Session, username: str, role: str) -> list:
    """Returns list of directions user can see. Admin/dcf_dir sees ALL."""
    if role in ("admin", "dcf_dir", "dcf_sub"):
        return ALL_DIRS
    u = db.query(User).filter_by(username=username).first()
    if not u: return []
    try:
        dirs = json.loads(u.directions or "[]")
        return dirs if dirs else ALL_DIRS
    except: return ALL_DIRS

def get_budget_status(db: Session, imputation: str, year: int, extra=0) -> str:
    if not imputation: return "OK"
    bl = db.query(BudgetLine).filter_by(imputation=imputation, year=year).first()
    if not bl: return "OK"
    used = db.query(sqlfunc.coalesce(sqlfunc.sum(Transaction.montant), 0))\
             .filter_by(imputation=imputation, year=year, status="validated").scalar() or 0
    return "OK" if (bl.budget_cp - used - extra) >= 0 else "NON"

# ── STARTUP ───────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app):
    try:
        Base.metadata.create_all(engine)
        log.info("DB tables ready")
    except Exception as e:
        log.error("create_all: %s", e)
    engine.dispose()
    db = SessionLocal()
    try:
        existing = db.query(User).filter_by(username=ADMIN_USER).first()
        if not existing:
            db.add(User(username=ADMIN_USER, password=_hash(ADMIN_PASS),
                        full_name="Administrateur", role="admin", is_active=True,
                        directions="[]"))
            db.commit()
            log.info("Admin created: %s", ADMIN_USER)
        elif existing.role != "admin":
            existing.role = "admin"; existing.is_active = True; db.commit()
    except Exception as e:
        db.rollback(); log.error("seed: %s", e)
    finally: db.close()
    yield

app = FastAPI(title="CAMTEL Budget v12", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ── PAGES ─────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def home(): return HTMLResponse(APP_HTML)

@app.get("/login", response_class=HTMLResponse)
def login_page(): return HTMLResponse(LOGIN_HTML)

# ── AUTH ──────────────────────────────────────────────────────────────
@app.post("/api/login")
def do_login(payload: dict, db: Session = Depends(get_db)):
    u = db.query(User).filter(User.username == payload.get("username",""),
                               User.is_active.is_(True)).first()
    if not u or u.password != _hash(payload.get("password","")):
        raise HTTPException(401, "Identifiants incorrects")
    dirs = json.loads(u.directions or "[]")
    token = SER.dumps({"u": u.username, "role": u.role, "name": u.full_name})
    resp = JSONResponse({"token": token, "role": u.role, "name": u.full_name, "directions": dirs})
    resp.set_cookie("session", token, httponly=True, samesite="lax",
                    secure=not IS_SQLITE, max_age=86400*7)
    return resp

@app.post("/api/logout")
def do_logout():
    r = JSONResponse({"ok": True}); r.delete_cookie("session"); return r

@app.get("/api/me")
def api_me(request: Request, db: Session = Depends(get_db)):
    tok = current_user(request)
    db_user = db.query(User).filter_by(username=tok["u"], is_active=True).first()
    if not db_user: raise HTTPException(401, "User not found")
    years = [r[0] for r in db.query(Transaction.year).distinct().order_by(Transaction.year.desc()).all()]
    if not years: years = [date.today().year]
    if date.today().year not in years: years.insert(0, date.today().year)
    dirs = get_user_directions(db, db_user.username, db_user.role)
    return {"u": db_user.username, "role": db_user.role, "name": db_user.full_name,
            "years": sorted(set(years), reverse=True), "directions": dirs}

# ── RESET ADMIN ───────────────────────────────────────────────────────
@app.get("/api/reset-admin")
def reset_admin(secret: str = "", db: Session = Depends(get_db)):
    if secret != "CAMTEL2025reset": raise HTTPException(403, "Wrong secret")
    new_pass = "Admin@2025!"
    u = db.query(User).filter_by(username=ADMIN_USER).first()
    if u:
        u.password = _hash(new_pass); u.is_active = True; u.role = "admin"; db.commit()
        return {"ok": True, "username": ADMIN_USER, "new_password": new_pass}
    db.add(User(username=ADMIN_USER, password=_hash(new_pass),
                full_name="Administrateur", role="admin", is_active=True, directions="[]"))
    db.commit()
    return {"ok": True, "username": ADMIN_USER, "new_password": new_pass, "created": True}

# ── DASHBOARD ─────────────────────────────────────────────────────────
@app.get("/api/dashboard")
def api_dashboard(year: int = 0, direction: str = "",
                  request: Request = None, db: Session = Depends(get_db)):
    tok = current_user(request)
    if not year: year = date.today().year
    user_dirs = get_user_directions(db, tok["u"], tok["role"])

    def filt_bl(q):
        q = q.filter(BudgetLine.year == year)
        if direction: q = q.filter(BudgetLine.direction == direction)
        elif tok["role"] not in ("admin","dcf_dir","dcf_sub"):
            q = q.filter(BudgetLine.direction.in_(user_dirs))
        return q

    def filt_tx(q, s="validated"):
        q = q.filter(Transaction.year == year)
        if s: q = q.filter(Transaction.status == s)
        if direction: q = q.filter(Transaction.direction == direction)
        elif tok["role"] not in ("admin","dcf_dir","dcf_sub"):
            q = q.filter(Transaction.direction.in_(user_dirs))
        return q

    budget  = filt_bl(db.query(sqlfunc.coalesce(sqlfunc.sum(BudgetLine.budget_cp), 0))).scalar() or 0
    engage  = filt_tx(db.query(sqlfunc.coalesce(sqlfunc.sum(Transaction.montant), 0))).scalar() or 0
    pending = filt_tx(db.query(sqlfunc.coalesce(sqlfunc.sum(Transaction.montant), 0)), "pending").scalar() or 0
    tx_count= filt_tx(db.query(sqlfunc.count(Transaction.id)), None).scalar() or 0

    monthly = []
    for m in range(1, 13):
        q = db.query(sqlfunc.coalesce(sqlfunc.sum(Transaction.montant), 0))\
              .filter(Transaction.year==year, Transaction.status=="validated",
                      Transaction.date_reception.like(f"{year}-{m:02d}%"))
        if direction: q = q.filter(Transaction.direction==direction)
        elif tok["role"] not in ("admin","dcf_dir","dcf_sub"):
            q = q.filter(Transaction.direction.in_(user_dirs))
        monthly.append({"m": m, "label": MONTHS_FR[m-1], "val": q.scalar() or 0})

    by_dir = []
    q_dirs = db.query(Transaction.direction, sqlfunc.sum(Transaction.montant))\
               .filter(Transaction.year==year, Transaction.status=="validated")\
               .group_by(Transaction.direction)
    if tok["role"] not in ("admin","dcf_dir","dcf_sub"):
        q_dirs = q_dirs.filter(Transaction.direction.in_(user_dirs))
    for d, amt in q_dirs.all():
        bl = db.query(sqlfunc.coalesce(sqlfunc.sum(BudgetLine.budget_cp), 0))\
               .filter_by(year=year, direction=d).scalar() or 0
        by_dir.append({"d": d, "eng": amt or 0, "bud": bl})

    return {"year": year, "budget": budget, "engage": engage, "pending": pending,
            "available": budget-engage, "pct": round(engage/budget*100,1) if budget else 0,
            "tx_count": tx_count, "monthly": monthly, "by_dir": by_dir}

# ── BUDGET LINES ──────────────────────────────────────────────────────
@app.get("/api/budget-lines")
def get_bls(year: int = 0, direction: str = "",
            request: Request = None, db: Session = Depends(get_db)):
    tok = current_user(request)
    if not year: year = date.today().year
    user_dirs = get_user_directions(db, tok["u"], tok["role"])
    q = db.query(BudgetLine).filter_by(year=year)
    if direction:
        q = q.filter_by(direction=direction)
    elif tok["role"] not in ("admin","dcf_dir","dcf_sub"):
        q = q.filter(BudgetLine.direction.in_(user_dirs))
    result = []
    for bl in q.order_by(BudgetLine.direction, BudgetLine.imputation).all():
        used = db.query(sqlfunc.coalesce(sqlfunc.sum(Transaction.montant), 0))\
                 .filter_by(imputation=bl.imputation, year=year, status="validated").scalar() or 0
        dispo = bl.budget_cp - used
        result.append({
            "id": bl.id, "year": bl.year, "direction": bl.direction,
            "imputation": bl.imputation, "libelle": bl.libelle,
            "nature": bl.nature, "budget_cp": bl.budget_cp,
            "budget_cp_fmt": fmt_amount(bl.budget_cp),
            "engage": used, "engage_fmt": fmt_amount(used),
            "dispo": dispo, "dispo_fmt": fmt_amount(abs(dispo)),
            "dispo_ok": dispo >= 0,
            "pct": round(used/bl.budget_cp*100,1) if bl.budget_cp else 0
        })
    return result

@app.post("/api/budget-lines")
def create_bl(data: dict, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] not in ("admin","dcf_dir","dcf_sub"): raise HTTPException(403)
    ex = db.query(BudgetLine).filter_by(year=data["year"],direction=data["direction"],
                                        imputation=data["imputation"]).first()
    if ex:
        ex.libelle=data.get("libelle",ex.libelle); ex.nature=data.get("nature",ex.nature)
        ex.budget_cp=float(data.get("budget_cp",ex.budget_cp)); db.commit()
        return {"id": ex.id, "updated": True}
    bl = BudgetLine(year=int(data["year"]), direction=data["direction"],
                    imputation=data["imputation"], libelle=data.get("libelle",""),
                    nature=data.get("nature","DEPENSE COURANTE"),
                    budget_cp=float(data.get("budget_cp",0)))
    db.add(bl); db.commit()
    return {"id": bl.id, "created": True}

@app.put("/api/budget-lines/{bid}")
def update_bl(bid: int, data: dict, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] not in ("admin","dcf_dir","dcf_sub"): raise HTTPException(403)
    bl = db.get(BudgetLine, bid)
    if not bl: raise HTTPException(404)
    for k in ("libelle","nature","budget_cp"):
        if k in data: setattr(bl, k, data[k])
    db.commit(); return {"ok": True}

@app.delete("/api/budget-lines/{bid}")
def delete_bl(bid: int, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] != "admin": raise HTTPException(403)
    bl = db.get(BudgetLine, bid)
    if not bl: raise HTTPException(404)
    db.delete(bl); db.commit(); return {"ok": True}

# ── TRANSACTIONS ──────────────────────────────────────────────────────
@app.get("/api/transactions")
def get_txs(year: int = 0, direction: str = "", page: int = 1, search: str = "",
            request: Request = None, db: Session = Depends(get_db)):
    tok = current_user(request)
    if not year: year = date.today().year
    user_dirs = get_user_directions(db, tok["u"], tok["role"])
    q = db.query(Transaction).filter_by(year=year)
    if direction: q = q.filter_by(direction=direction)
    elif tok["role"] not in ("admin","dcf_dir","dcf_sub"):
        q = q.filter(Transaction.direction.in_(user_dirs))
    if search:
        s = f"%{search}%"
        q = q.filter(Transaction.intitule.ilike(s) | Transaction.imputation.ilike(s))
    total = q.count(); PER = 50
    rows  = q.order_by(Transaction.date_reception.desc()).offset((page-1)*PER).limit(PER).all()
    return {"total": total, "page": page, "per_page": PER,
            "rows": [{
                "id": r.id, "date": r.date_reception, "direction": r.direction,
                "imputation": r.imputation, "intitule": r.intitule, "libelle": r.libelle,
                "objet": r.objet, "nature": r.nature, "designation": r.designation,
                "montant": r.montant, "montant_fmt": fmt_amount(r.montant),
                "status": r.status, "statut_budget": r.statut_budget,
                "created_by": r.created_by, "created_by_name": r.created_by_name,
                "created_at": r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else ""
            } for r in rows]}

@app.post("/api/transactions")
def create_tx(data: dict, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] not in ("admin","dcf_dir","dcf_sub","agent_plus","agent"): raise HTTPException(403)
    db_user = db.query(User).filter_by(username=u["u"]).first()
    montant = float(data.get("montant",0))
    imp = data.get("imputation",""); yr = int(data.get("year", date.today().year))
    tx = Transaction(
        date_reception=data.get("date",date.today().isoformat()),
        direction=data.get("direction",""), imputation=imp,
        intitule=data.get("intitule",""), libelle=data.get("libelle",""),
        objet=data.get("objet",""), nature=data.get("nature","DEPENSE COURANTE"),
        designation=data.get("designation",""), code_ref=data.get("code_ref",""),
        montant=montant, year=yr, status=data.get("status","validated"),
        statut_budget=get_budget_status(db,imp,yr,montant),
        created_by=u["u"], created_by_name=db_user.full_name if db_user else u["u"]
    )
    db.add(tx); db.commit()
    return {"id": tx.id, "statut_budget": tx.statut_budget}

@app.put("/api/transactions/{tid}")
def update_tx(tid: int, data: dict, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    tx = db.get(Transaction, tid)
    if not tx: raise HTTPException(404)
    if u["role"] not in ("admin","dcf_dir","dcf_sub") and tx.created_by != u["u"]: raise HTTPException(403)
    for k in ("date_reception","direction","imputation","intitule","libelle","objet","nature","designation","montant","status"):
        if k in data: setattr(tx, k, data[k])
    tx.statut_budget = get_budget_status(db, tx.imputation, tx.year, 0)
    db.commit(); return {"ok": True}

@app.delete("/api/transactions/{tid}")
def delete_tx(tid: int, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    tx = db.get(Transaction, tid)
    if not tx: raise HTTPException(404)
    if u["role"] not in ("admin","dcf_dir") and tx.created_by != u["u"]: raise HTTPException(403)
    db.delete(tx); db.commit(); return {"ok": True}

# ── IMPORT TRANSACTIONS ───────────────────────────────────────────────
@app.post("/api/import/transactions")
async def import_txs(request: Request, file: UploadFile = File(...),
                     year: int = Form(...), db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] not in ("admin","dcf_dir","dcf_sub","agent_plus"): raise HTTPException(403)
    db_user = db.query(User).filter_by(username=u["u"]).first()
    uname = db_user.full_name if db_user else u["u"]
    raw = await file.read()
    created = 0; errors = []
    try:
        all_rows = read_file_rows(raw, file.filename or "")
        hi = find_header_row(all_rows)
        for ri, row in enumerate(all_rows[hi+1:], hi+2):
            try:
                r = parse_camtel_tx(row)
                if not r: continue
                tx = Transaction(
                    date_reception=r["date"], direction=r["direction"],
                    imputation=r["imputation"], intitule=r["intitule"],
                    libelle=r["libelle"], nature=r["nature"],
                    designation=r["designation"], code_ref=r["code_ref"],
                    montant=r["montant"], year=year, status="validated",
                    statut_budget=get_budget_status(db,r["imputation"],year),
                    created_by=u["u"], created_by_name=uname
                )
                db.add(tx); created += 1
            except Exception as e:
                errors.append(f"L{ri}: {e}")
        db.commit()
    except Exception as e:
        db.rollback(); return JSONResponse({"created":0,"errors":[str(e)]}, status_code=400)
    return {"created": created, "errors": errors[:5]}

# ── IMPORT BUDGET LINES ───────────────────────────────────────────────
@app.post("/api/import/budget-lines")
async def import_bls(request: Request, file: UploadFile = File(...),
                     year: int = Form(0), db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] not in ("admin","dcf_dir","dcf_sub"): raise HTTPException(403)
    raw = await file.read()
    created = updated = skipped = 0; errors = []
    try:
        all_rows = read_file_rows(raw, file.filename or "")
        if not all_rows: return {"created":0,"updated":0,"skipped":0,"errors":["Empty file"]}
        header_str = " ".join(str(c) for c in all_rows[0]).upper()
        is_standard = "YEAR" in header_str or "BUDGET CP" in header_str or "ANNEE" in header_str

        if is_standard:
            headers = [str(c).strip().upper() for c in all_rows[0]]
            def gcol(row, *keys):
                for k in keys:
                    for i, h in enumerate(headers):
                        if k.upper() in h and i < len(row):
                            v = str(row[i]).strip()
                            if v and v.upper() not in ("NONE",""): return v
                return ""
            for ri, row in enumerate(all_rows[1:], 2):
                try:
                    yr_s = gcol(row,"YEAR","ANNEE") or str(year)
                    dirn = gcol(row,"DIRECTION").upper()
                    imp  = gcol(row,"IMPUTATION")
                    lib  = gcol(row,"LIBELLE","DESCRIPTION")
                    nat  = gcol(row,"NATURE") or "DEPENSE COURANTE"
                    bcp_s= gcol(row,"BUDGET CP","MONTANT","BUDGET")
                    if not dirn or not imp: skipped+=1; continue
                    yr2 = int(float(yr_s)) if yr_s else (year or 2025)
                    bcp = clean_amount(bcp_s)
                    ex = db.query(BudgetLine).filter_by(year=yr2,direction=dirn,imputation=imp).first()
                    if ex: ex.libelle=lib;ex.nature=nat;ex.budget_cp=bcp;updated+=1
                    else:
                        db.add(BudgetLine(year=yr2,direction=dirn,imputation=imp,
                                          libelle=lib,nature=nat,budget_cp=bcp));created+=1
                except Exception as e: errors.append(f"L{ri}: {e}")
        else:
            target_year = year or 2025
            col_yr_map = {2024:17, 2025:18, 2026:19}
            bcp_col = col_yr_map.get(target_year, 18)
            for ri, row in enumerate(all_rows, 1):
                try:
                    if len(row) < 12: skipped+=1; continue
                    dirn = str(row[3]).strip().upper()
                    imp  = str(row[10]).strip()
                    lib  = str(row[11]).strip()
                    if not dirn or not imp or not imp[0:1].isdigit(): skipped+=1; continue
                    bcp_s = str(row[bcp_col]).strip() if len(row)>bcp_col else "0"
                    bcp = clean_amount(bcp_s)
                    nat = "DEPENSE DE CAPITAL" if any(k in lib.upper() for k in ("CAPITAL","INVESTIS")) \
                          else "DEPENSE COURANTE"
                    ex = db.query(BudgetLine).filter_by(year=target_year,direction=dirn,imputation=imp).first()
                    if ex: ex.libelle=lib;ex.nature=nat;ex.budget_cp=bcp;updated+=1
                    else:
                        db.add(BudgetLine(year=target_year,direction=dirn,imputation=imp,
                                          libelle=lib,nature=nat,budget_cp=bcp));created+=1
                except Exception as e: errors.append(f"L{ri}: {e}")
        db.commit()
    except Exception as e:
        db.rollback(); return JSONResponse({"created":0,"updated":0,"errors":[str(e)]}, status_code=400)
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors[:5]}

# ── EXPORTS ───────────────────────────────────────────────────────────
@app.get("/api/export/transactions")
def export_txs(year: int = 0, direction: str = "",
               request: Request = None, db: Session = Depends(get_db)):
    current_user(request)
    if not year: year = date.today().year
    q = db.query(Transaction).filter_by(year=year)
    if direction: q = q.filter_by(direction=direction)
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(["DATE","DIRECTION","INTITULE","LIBELLE","OBJET","NATURE",
                "DESIGNATION","IMPUTATION","MONTANT","STATUT_BUDGET","INITIATEUR"])
    for r in q.order_by(Transaction.date_reception).all():
        w.writerow([r.date_reception,r.direction,r.intitule,r.libelle,r.objet,
                    r.nature,r.designation,r.imputation,r.montant,
                    r.statut_budget,r.created_by_name])
    return StreamingResponse(iter([buf.getvalue().encode("utf-8-sig")]),
                             media_type="text/csv",
                             headers={"Content-Disposition":f"attachment;filename=transactions_{year}.csv"})

@app.get("/api/export/template-transactions")
def tpl_tx():
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Transactions"
        hdr_fill = PatternFill("solid", fgColor="1a3a6b")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        headers = ["DATE ENGAGEMENT","DIRECTION","INTITULE DE LA COMMANDE","LIBELLE",
                   "NATURE DE LA DEPENSE (DEPENSE COURANTE, DEPENSE DE CAPITAL)",
                   "DESIGNATION (OM, MM, MT, NC, BC, LC, ADDB)","IMPUTATION COMPTABLE","MONTANT"]
        # Title row
        ws.append(["SITUATION DES ENGAGEMENTS BUDGETAIRES DU MOIS DE ________"] + [""]*7)
        ws.merge_cells("A1:H1")
        ws["A1"].font = Font(bold=True, size=11, color="1a3a6b")
        ws["A1"].alignment = Alignment(horizontal="center")
        # Header row
        ws.append(headers)
        for cell in ws[2]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        # Sample rows
        samples = [
            ["06/01/2025","DRH","COTISATION CNPS DECEMBRE 2024","PENSION VIEILLESSE",
             "DEPENSE COURANTE","NC","SP4/DRH/AD0002/VD0029/T00362/66410200>>>PENSION VIEILLESSE","291959762"],
            ["07/01/2025","DCF","FACTURE ELECTRICITE","FOURNISSEURS",
             "DEPENSE COURANTE","","SP4/DCF/AD0028/VD0060/PD0006/90530000>>>FOURNISSEURS","188225748"],
            ["15/03/2025","DSPI","ACQUISITION EQUIPEMENT","EQUIPEMENTS INFORMATIQUES",
             "DEPENSE DE CAPITAL","MT","SP4/DSPI/AD0010/VD0020/22000000>>>EQUIPEMENTS","50000000"],
        ]
        for s in samples: ws.append(s)
        # Column widths
        widths = [14,12,40,30,30,15,50,15]
        for i,w in enumerate(widths,1):
            ws.column_dimensions[chr(64+i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(iter([buf.read()]),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition":"attachment;filename=template_transactions.xlsx"})
    # CSV fallback
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(["SITUATION DES ENGAGEMENTS BUDGETAIRES","","","","","","",""])
    w.writerow(["DATE ENGAGEMENT","DIRECTION","INTITULE DE LA COMMANDE","LIBELLE",
                "NATURE DE LA DEPENSE","DESIGNATION","IMPUTATION COMPTABLE","MONTANT"])
    w.writerow(["06/01/2025","DRH","COTISATION CNPS","PENSION VIEILLESSE","DEPENSE COURANTE",
                "NC","SP4/DRH/AD0002/66410200>>>PENSION VIEILLESSE","291959762"])
    return StreamingResponse(iter([buf.getvalue().encode("utf-8-sig")]),
                             media_type="text/csv",
                             headers={"Content-Disposition":"attachment;filename=template_transactions.csv"})

@app.get("/api/export/template-budget-lines")
def tpl_bl(year: int = 2026):
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = f"Budget {year}"
        hdr_fill = PatternFill("solid", fgColor="1a3a6b")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        ws.append([f"CAMTEL — BUDGET CP {year} — MODÈLE D'IMPORTATION"] + [""]*5)
        ws.merge_cells("A1:F1")
        ws["A1"].font = Font(bold=True, size=12, color="1a3a6b")
        ws["A1"].alignment = Alignment(horizontal="center")
        headers = ["YEAR","DIRECTION","IMPUTATION COMPTABLE","LIBELLE","NATURE","BUDGET CP (FCFA)"]
        ws.append(headers)
        for cell in ws[2]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        samples = [
            [year,"DCF","60410000","FOURNITURES DE BUREAU","DEPENSE COURANTE",5000000],
            [year,"DRH","66410200","CHARGES SOCIALES CNPS","DEPENSE COURANTE",800000000],
            [year,"DSPI","22000000","ACQUISITION EQUIPEMENTS","DEPENSE DE CAPITAL",500000000],
            [year,"DAP","62210000","ENTRETIEN VEHICULES","DEPENSE COURANTE",30000000],
            [year,"DG","63110000","HONORAIRES CONSEIL","DEPENSE COURANTE",50000000],
        ]
        for s in samples: ws.append(s)
        # Note row
        ws.append([""])
        ws.append(["NOTE: Remplacez les exemples par vos vraies données. "
                   "DIRECTION doit être une des directions CAMTEL officielles."])
        widths = [8,12,22,40,22,20]
        for i,w in enumerate(widths,1):
            ws.column_dimensions[chr(64+i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(iter([buf.read()]),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition":f"attachment;filename=template_budget_{year}.xlsx"})
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(["YEAR","DIRECTION","IMPUTATION COMPTABLE","LIBELLE","NATURE","BUDGET CP (FCFA)"])
    for s in [[year,"DCF","60410000","FOURNITURES","DEPENSE COURANTE",5000000],
              [year,"DRH","66410200","CHARGES SOCIALES","DEPENSE COURANTE",800000000]]:
        w.writerow(s)
    return StreamingResponse(iter([buf.getvalue().encode("utf-8-sig")]),
                             media_type="text/csv",
                             headers={"Content-Disposition":f"attachment;filename=template_budget_{year}.csv"})

# ── USERS ─────────────────────────────────────────────────────────────
@app.get("/api/users")
def get_users(request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] != "admin": raise HTTPException(403)
    return [{"id":x.id,"username":x.username,"full_name":x.full_name,
             "role":x.role,"email":x.email,"is_active":x.is_active,
             "directions": json.loads(x.directions or "[]")}
            for x in db.query(User).order_by(User.username).all()]

@app.post("/api/users")
def create_user(data: dict, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] != "admin": raise HTTPException(403)
    if db.query(User).filter_by(username=data["username"]).first():
        raise HTTPException(400, "Username exists")
    dirs = json.dumps(data.get("directions", []))
    nu = User(username=data["username"], password=_hash(data["password"]),
              full_name=data.get("full_name",""), role=data.get("role","agent"),
              email=data.get("email",""), directions=dirs)
    db.add(nu); db.commit(); return {"id": nu.id}

@app.put("/api/users/{uid}")
def update_user(uid: int, data: dict, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] != "admin": raise HTTPException(403)
    usr = db.get(User, uid)
    if not usr: raise HTTPException(404)
    for k in ("full_name","role","email","is_active"):
        if k in data: setattr(usr, k, data[k])
    if "directions" in data: usr.directions = json.dumps(data["directions"])
    if data.get("password"): usr.password = _hash(data["password"])
    db.commit(); return {"ok": True}

@app.delete("/api/users/{uid}")
def delete_user(uid: int, request: Request = None, db: Session = Depends(get_db)):
    u = current_user(request)
    if u["role"] != "admin": raise HTTPException(403)
    usr = db.get(User, uid)
    if not usr: raise HTTPException(404)
    db.delete(usr); db.commit(); return {"ok": True}

@app.get("/api/directions")
def get_directions(request: Request = None, db: Session = Depends(get_db)):
    tok = current_user(request)
    return get_user_directions(db, tok["u"], tok["role"])

@app.get("/version")
def version(): return {"version": "v12", "status": "ok"}

# ══════════════════════════════════════════════════════════════════════
# LOGIN PAGE
# ══════════════════════════════════════════════════════════════════════
LOGIN_HTML = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>CAMTEL — Connexion</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',sans-serif;background:#020a18;min-height:100vh;
  display:flex;align-items:center;justify-content:center;
  background-image:radial-gradient(ellipse at 30% 40%,rgba(0,80,180,.4),transparent 60%),
    radial-gradient(ellipse at 75% 70%,rgba(0,40,120,.3),transparent 55%)}
.box{background:rgba(6,16,44,.95);border:1px solid rgba(0,150,255,.2);border-radius:16px;
  padding:44px 40px;width:min(420px,94vw);box-shadow:0 24px 64px rgba(0,0,0,.6)}
.logo{text-align:center;margin-bottom:32px}
.logo-icon{font-size:40px;margin-bottom:10px}
.logo h1{font-size:24px;font-weight:800;color:#fff;letter-spacing:2px}
.logo p{font-size:11px;color:rgba(0,210,255,.65);margin-top:5px;text-transform:uppercase;letter-spacing:.12em}
label{display:block;font-size:11px;font-weight:700;color:#6a9cc0;text-transform:uppercase;
  letter-spacing:.07em;margin-bottom:5px;margin-top:18px}
input{width:100%;padding:12px 14px;border-radius:8px;border:1.5px solid rgba(0,120,200,.3);
  background:rgba(4,12,36,.95);color:#d0e8ff;font-size:14px;font-family:inherit;transition:border .2s}
input:focus{outline:none;border-color:#00d4ff;box-shadow:0 0 0 3px rgba(0,180,255,.1)}
.btn{width:100%;margin-top:24px;padding:14px;border-radius:9px;border:none;
  background:linear-gradient(135deg,#0d3a7a,#1560af);color:#fff;font-size:14px;
  font-weight:700;cursor:pointer;transition:all .2s}
.btn:hover{background:linear-gradient(135deg,#1550a0,#00aadd);transform:translateY(-1px)}
.btn:disabled{opacity:.6;cursor:not-allowed;transform:none}
#err{min-height:22px;margin-top:14px;text-align:center;font-size:12px;color:#ff7070;font-weight:600}
.hint{text-align:center;font-size:10px;color:rgba(80,120,160,.4);margin-top:20px}
</style>
</head>
<body>
<div class="box">
  <div class="logo">
    <div class="logo-icon">🌐</div>
    <h1>CAMTEL</h1>
    <p>Système de Gestion Budgétaire</p>
  </div>
  <label>Identifiant</label>
  <input type="text" id="user" placeholder="admin" autocomplete="username" required>
  <label>Mot de passe</label>
  <input type="password" id="pass" placeholder="••••••••" autocomplete="current-password"
         required onkeydown="if(event.key==='Enter')doLogin()">
  <button class="btn" id="btn" onclick="doLogin()">Connexion</button>
  <div id="err"></div>
  <div class="hint">CAMTEL Budget v12</div>
</div>
<script>
async function doLogin() {
  const btn=document.getElementById('btn'),err=document.getElementById('err');
  const u=document.getElementById('user').value.trim(),p=document.getElementById('pass').value;
  if(!u||!p){err.textContent='Remplissez tous les champs';return}
  btn.disabled=true;btn.textContent='...';err.textContent='';
  try{
    const r=await fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:u,password:p})});
    const d=await r.json();
    if(r.ok){sessionStorage.setItem('tok',d.token);sessionStorage.setItem('role',d.role);sessionStorage.setItem('uname',d.name||u);window.location='/';}
    else{err.textContent='❌ '+(d.detail||'Identifiants incorrects');btn.disabled=false;btn.textContent='Connexion';}
  }catch(e){err.textContent='❌ Erreur: '+e.message;btn.disabled=false;btn.textContent='Connexion';}
}
</script>
</body>
</html>"""

# ══════════════════════════════════════════════════════════════════════
# MAIN APP HTML
# ══════════════════════════════════════════════════════════════════════
APP_HTML = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>CAMTEL Budget</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{--navy:#020a18;--card:#06102c;--bdr:rgba(0,140,220,.18);--txt:#ddeeff;--muted:#6a9cc0}
body{font-family:'Segoe UI',sans-serif;background:var(--navy);color:var(--txt);font-size:13px}
body::before{content:"";position:fixed;inset:0;z-index:-1;pointer-events:none;
  background:radial-gradient(ellipse at 18% 22%,rgba(0,60,150,.3),transparent 52%),
    radial-gradient(ellipse at 82% 78%,rgba(0,30,100,.25),transparent 50%)}
header{background:linear-gradient(90deg,#020810,#091a44,#020810);border-bottom:1px solid var(--bdr);
  padding:0 14px;height:52px;display:flex;align-items:center;gap:8px;position:sticky;top:0;z-index:300;
  box-shadow:0 2px 20px rgba(0,0,0,.5)}
.brand h1{font-size:14px;font-weight:800;color:#fff;white-space:nowrap;letter-spacing:1px}
.brand p{font-size:9px;color:rgba(0,200,255,.6);text-transform:uppercase;letter-spacing:.1em}
nav{display:flex;gap:2px;margin:0 8px;flex:1;overflow-x:auto}
nav::-webkit-scrollbar{height:2px}
nav button{background:rgba(0,50,120,.2);color:rgba(170,210,255,.7);border:1px solid rgba(0,110,220,.12);
  padding:5px 12px;border-radius:5px;cursor:pointer;font-size:11px;font-weight:600;white-space:nowrap;transition:all .18s}
nav button:hover{background:rgba(0,90,200,.35);color:#fff}
nav button.on{background:rgba(0,130,255,.25);color:#00d4ff;border-color:rgba(0,190,255,.4)}
.upill{background:rgba(255,255,255,.07);padding:3px 10px;border-radius:20px;font-size:11px;white-space:nowrap}
.btn-out{background:#b71c1c;color:#fff;border:none;padding:5px 11px;border-radius:5px;cursor:pointer;font-size:11px;font-weight:700}
.wrap{padding:12px 14px}
.toolbar{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:12px;padding:8px 12px;
  background:rgba(0,8,28,.65);border:1px solid rgba(0,90,190,.14);border-radius:8px}
.toolbar label{font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase}
.toolbar select,.toolbar input[type=text]{background:rgba(4,10,36,.9);color:#c0d8f0;
  border:1px solid rgba(0,90,200,.28);border-radius:6px;padding:5px 8px;font-size:12px;font-family:inherit}
.tab{display:none}.tab.on{display:block}
/* KPI */
.krow{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:14px}
@media(max-width:700px){.krow{grid-template-columns:1fr 1fr}}
.kpi{background:linear-gradient(135deg,rgba(6,18,52,.95),rgba(10,26,68,.9));
  border:1px solid var(--bdr);border-radius:12px;padding:14px 16px}
.kpi .kl{font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted);font-weight:700}
.kpi .kv{font-size:17px;font-weight:800;margin-top:4px;letter-spacing:.5px}
.kpi .ks{font-size:10px;color:#4a7090;margin-top:2px}
.kb .kv{color:#1e88e5}.kr .kv{color:#e53935}.ky .kv{color:#fb8c00}.kg .kv{color:#43a047}
/* CARD */
.card{background:rgba(6,14,44,.9);border:1px solid var(--bdr);border-radius:12px;overflow:hidden;margin-bottom:12px}
.ch{padding:10px 14px;border-bottom:1px solid var(--bdr);
  background:linear-gradient(90deg,rgba(0,20,70,.9),rgba(0,12,44,.95));
  display:flex;align-items:center;justify-content:space-between}
.ch h2{font-size:12px;font-weight:700;color:#7ab8d8}
.cb{padding:12px}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
@media(max-width:700px){.g2{grid-template-columns:1fr}}
/* BUTTONS */
.btn{padding:6px 13px;border-radius:7px;border:none;cursor:pointer;font-size:11px;font-weight:700;
  font-family:inherit;transition:all .15s;display:inline-flex;align-items:center;gap:4px}
.bp{background:linear-gradient(135deg,#0d3a7a,#1560af);color:#fff}.bp:hover{background:#0090cc}
.bs{background:#1e2d45;color:#8aaccc;border:1px solid #2a3d5a}.bs:hover{background:#263650}
.bd{background:#b71c1c;color:#fff}.bd:hover{background:#d32f2f}
.bg{background:rgba(0,160,70,.2);color:#4caf50;border:1px solid rgba(0,160,70,.3)}
.bsm{padding:3px 8px;font-size:10px;border-radius:5px}
/* TABLE */
.tw{overflow-x:auto;border-radius:8px;border:1px solid var(--bdr)}
table{width:100%;border-collapse:collapse;font-size:11px}
th{background:rgba(0,18,54,.9);color:var(--muted);font-weight:700;text-transform:uppercase;
  letter-spacing:.04em;padding:8px 10px;text-align:left;position:sticky;top:0;white-space:nowrap}
td{padding:7px 10px;border-bottom:1px solid rgba(0,50,110,.15);color:#bbd4f0;vertical-align:middle}
tr:hover td{background:rgba(0,50,130,.12)}
.sel-row{background:rgba(0,100,200,.15)!important}
.sel-row td{background:rgba(0,100,200,.15)!important}
/* BADGES */
.bok{background:rgba(0,160,70,.14);color:#4caf50;border:1px solid rgba(0,160,70,.3);
  padding:2px 8px;border-radius:12px;font-size:10px;font-weight:700}
.bnon{background:rgba(220,30,30,.14);color:#ef5350;border:1px solid rgba(220,30,30,.3);
  padding:2px 8px;border-radius:12px;font-size:10px;font-weight:700}
.bpend{background:rgba(255,160,0,.14);color:#fb8c00;border:1px solid rgba(255,160,0,.3);
  padding:2px 8px;border-radius:12px;font-size:10px;font-weight:700}
/* MODAL */
.mbg{display:none;position:fixed;inset:0;background:rgba(0,6,28,.65);
  z-index:500;align-items:flex-start;justify-content:center;padding-top:30px;overflow-y:auto}
.mbg.open{display:flex}
.modal{background:rgba(4,12,40,.97);border:1px solid rgba(0,130,255,.22);border-radius:14px;
  width:min(640px,96vw);max-height:92vh;overflow-y:auto;box-shadow:0 24px 64px rgba(0,0,0,.6)}
.mh{padding:12px 16px;border-bottom:1px solid var(--bdr);display:flex;align-items:center;
  justify-content:space-between;background:rgba(0,12,40,.8);position:sticky;top:0;z-index:10}
.mh h3{font-size:13px;font-weight:700}
.mb{padding:16px}.mf{padding:10px 16px;border-top:1px solid var(--bdr);
  display:flex;gap:8px;justify-content:flex-end;background:rgba(0,8,30,.7)}
.fc2{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.fc3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}
/* FORM */
.fld{margin-bottom:10px}
.fld label{display:block;font-size:10px;font-weight:700;color:var(--muted);
  text-transform:uppercase;letter-spacing:.04em;margin-bottom:4px}
.fld input,.fld select,.fld textarea{width:100%;padding:8px 10px;border-radius:7px;
  border:1.5px solid rgba(0,100,200,.22);background:rgba(4,10,36,.95);
  color:#c0d8f0;font-size:12px;font-family:inherit}
.fld input:focus,.fld select:focus,.fld textarea:focus{outline:none;border-color:#00d4ff}
/* ALERTS */
.alrt{padding:8px 12px;border-radius:7px;font-size:11px;margin-bottom:8px;border-left:3px solid}
.ar{background:rgba(180,20,20,.12);color:#ef5350;border-color:#e53935}
.ag{background:rgba(0,150,60,.1);color:#4caf50;border-color:#43a047}
.ay{background:rgba(200,130,0,.1);color:#fb8c00;border-color:#fb8c00}
.ab{background:rgba(0,90,220,.1);color:#64b5f6;border-color:#1976d2}
/* PROGRESS */
.prg{height:5px;background:rgba(0,30,80,.5);border-radius:3px;overflow:hidden;margin-top:4px}
.prf{height:100%;border-radius:3px;transition:width .4s}
/* DROP ZONE */
.dz{display:block;border:2px dashed rgba(0,140,255,.3);border-radius:10px;padding:24px 16px;
  text-align:center;cursor:pointer;background:rgba(0,30,80,.08);transition:all .2s;margin-bottom:10px}
.dz:hover,.dz.over{border-color:rgba(0,200,255,.7);background:rgba(0,50,120,.18)}
.dz .di{font-size:36px;margin-bottom:6px;pointer-events:none}
.dz .dl{font-size:13px;font-weight:700;color:#7ab8d8;pointer-events:none}
.dz .ds{font-size:10px;color:#3a6888;margin-top:4px;pointer-events:none}
.dz .dn{margin-top:6px;font-size:12px;color:#00d4ff;font-weight:700;min-height:16px;pointer-events:none}
/* DIRECTIONS TAGS */
.dir-tags{display:flex;flex-wrap:wrap;gap:4px;max-height:120px;overflow-y:auto;padding:4px;
  background:rgba(0,10,36,.5);border:1px solid rgba(0,80,180,.2);border-radius:6px}
.dir-tag{padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;cursor:pointer;
  background:rgba(0,50,120,.3);color:#7ab8d8;border:1px solid rgba(0,100,200,.2);transition:all .15s}
.dir-tag.sel{background:rgba(0,130,255,.3);color:#00d4ff;border-color:rgba(0,190,255,.5)}
/* AMOUNT — right aligned, no overflow */
.amt{text-align:right;white-space:nowrap;font-family:'Courier New',monospace;font-weight:600;font-size:11px}
/* PRINT STYLES */
@media print {
  body{background:#fff!important;color:#000!important}
  header,nav,.toolbar,.btn,.tw,.no-print{display:none!important}
  .print-area{display:block!important}
}
.print-area{display:none}
</style>
</head>
<body>

<header>
  <div class="brand"><h1>🌐 CAMTEL</h1><p>Budget v12</p></div>
  <nav>
    <button onclick="show('dashboard')" id="nb-dashboard" class="on">📊 Dashboard</button>
    <button onclick="show('transactions')" id="nb-transactions">📋 Transactions</button>
    <button onclick="show('budget')" id="nb-budget">💰 Budget</button>
    <button onclick="show('import')" id="nb-import">⬆ Import</button>
    <button onclick="show('users')" id="nb-users" style="display:none">👥 Utilisateurs</button>
  </nav>
  <div class="upill"><span id="h-name">—</span> <span id="h-role" style="opacity:.55"></span></div>
  <button class="btn-out" onclick="logout()">⏻</button>
</header>

<div class="wrap">
<div class="toolbar">
  <label>Année</label>
  <select id="g-yr" onchange="onFilters()"><option>2025</option></select>
  <label style="margin-left:10px">Direction</label>
  <select id="g-dir" onchange="onFilters()"><option value="">Toutes</option></select>
  <button class="btn bs" onclick="reload()" style="margin-left:auto">🔄</button>
</div>

<!-- ═══ DASHBOARD ═══ -->
<div id="tab-dashboard" class="tab on">
  <div class="krow">
    <div class="kpi kb"><div class="kl">Budget CP</div><div class="kv" id="k-bud">—</div><div class="ks" id="k-bud2"></div></div>
    <div class="kpi kr"><div class="kl">Engagé</div><div class="kv" id="k-eng">—</div><div class="ks" id="k-eng2"></div></div>
    <div class="kpi ky"><div class="kl">En attente</div><div class="kv" id="k-pend">—</div><div class="ks"></div></div>
    <div class="kpi kg"><div class="kl">Disponible</div><div class="kv" id="k-dispo">—</div><div class="ks" id="k-dispo2"></div></div>
  </div>
  <div class="g2">
    <div class="card">
      <div class="ch"><h2>📈 Engagements mensuels</h2></div>
      <div class="cb" style="padding:8px"><canvas id="ch-monthly" height="160" style="width:100%;display:block"></canvas></div>
    </div>
    <div class="card">
      <div class="ch"><h2>🏢 Par direction</h2></div>
      <div class="cb" style="padding:0;max-height:240px;overflow-y:auto" id="by-dir"></div>
    </div>
  </div>
</div>

<!-- ═══ TRANSACTIONS ═══ -->
<div id="tab-transactions" class="tab">
  <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center;flex-wrap:wrap">
    <input type="text" id="tx-q" placeholder="🔍 Rechercher..." oninput="txSearch()"
      style="background:rgba(4,10,36,.9);color:#c0d8f0;border:1px solid rgba(0,90,200,.28);
             border-radius:6px;padding:6px 10px;font-size:12px;width:200px">
    <button class="btn bp" onclick="openTxModal()">＋ Ajouter</button>
    <button class="btn bs" onclick="printSelected()" id="btn-print" style="display:none">🖨 Imprimer sélection</button>
    <button class="btn bs" onclick="window.open('/api/export/transactions?year='+yr()+'&direction='+dir())">⬇ CSV</button>
    <span id="tx-info" style="margin-left:auto;font-size:11px;color:var(--muted)"></span>
  </div>
  <div class="tw">
    <table>
      <thead><tr>
        <th style="width:32px"><input type="checkbox" id="chk-all" onchange="toggleAll(this)"></th>
        <th>Date</th><th>Direction</th><th>Libellé(s)</th>
        <th>Imputation</th><th>Nature</th><th class="amt">Montant (FCFA)</th>
        <th>Disponible</th><th>Initiateur</th><th></th>
      </tr></thead>
      <tbody id="tx-body"><tr><td colspan="10" style="text-align:center;padding:28px;color:#3a6888">Chargement...</td></tr></tbody>
    </table>
  </div>
  <div id="tx-pages" style="display:flex;gap:5px;margin-top:10px;flex-wrap:wrap"></div>
</div>

<!-- ═══ BUDGET ═══ -->
<div id="tab-budget" class="tab">
  <div style="display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap;align-items:center">
    <button class="btn bp" id="btn-add-bl" onclick="openBLModal()" style="display:none">＋ Ajouter</button>
    <button class="btn bs" onclick="window.open('/api/export/template-budget-lines?year='+yr())">⬇ Template Budget</button>
    <span id="bl-info" style="margin-left:auto;font-size:11px;color:var(--muted)"></span>
  </div>
  <div class="tw">
    <table>
      <thead><tr>
        <th>Direction</th><th>Imputation</th><th>Libellé(s)</th><th>Nature</th>
        <th class="amt">Budget CP (FCFA)</th><th class="amt">Engagé</th>
        <th class="amt">Disponible</th><th>%</th><th>Avis DCF</th><th></th>
      </tr></thead>
      <tbody id="bl-body"><tr><td colspan="10" style="text-align:center;padding:28px;color:#3a6888">Chargement...</td></tr></tbody>
    </table>
  </div>
</div>

<!-- ═══ IMPORT ═══ -->
<div id="tab-import" class="tab">
  <div class="g2">
    <div class="card">
      <div class="ch" style="background:linear-gradient(90deg,rgba(0,50,15,.9),rgba(0,30,8,.95))">
        <h2 style="color:#66bb6a">⬆ Importer Transactions</h2>
      </div>
      <div class="cb">
        <div class="alrt ab" style="font-size:10px;margin-bottom:12px">
          <b>Format CAMTEL :</b> DATE ENGAGEMENT · DIRECTION · INTITULE · LIBELLE · NATURE · DESIGNATION · IMPUTATION COMPTABLE · MONTANT
        </div>
        <div class="fld"><label>Année *</label><input type="number" id="tx-yr-imp" value="2025" min="2020" max="2035"></div>
        <input type="file" id="tx-file-inp" accept=".csv,.xlsx,.xls,.txt"
          style="position:fixed;left:-9999px;top:-9999px;opacity:0;width:1px;height:1px"
          onchange="onFile(this,'tx-file-name','tx-dz')">
        <label for="tx-file-inp" id="tx-dz" class="dz">
          <div class="di">📂</div>
          <div class="dl">Cliquer ici pour choisir un fichier</div>
          <div class="ds">CSV, XLSX, XLS acceptés</div>
          <div class="dn" id="tx-file-name"></div>
        </label>
        <div style="display:flex;gap:8px">
          <button class="btn bp" onclick="doImportTx()">⬆ Importer</button>
          <button class="btn bs" onclick="window.open('/api/export/template-transactions')">⬇ Template XLSX</button>
        </div>
        <div id="tx-imp-res" style="margin-top:10px"></div>
      </div>
    </div>
    <div class="card">
      <div class="ch" style="background:linear-gradient(90deg,rgba(0,25,75,.9),rgba(0,15,50,.95))">
        <h2 style="color:#42a5f5">⬆ Importer Lignes Budget</h2>
      </div>
      <div class="cb">
        <div class="alrt ab" style="font-size:10px;margin-bottom:12px">
          <b>Colonnes :</b> YEAR · DIRECTION · IMPUTATION COMPTABLE · LIBELLE · NATURE · BUDGET CP (FCFA)
        </div>
        <div class="fld"><label>Année (si absente du fichier)</label><input type="number" id="bl-yr-imp" value="2025" min="2020" max="2035"></div>
        <input type="file" id="bl-file-inp" accept=".csv,.xlsx,.xls,.txt"
          style="position:fixed;left:-9999px;top:-9999px;opacity:0;width:1px;height:1px"
          onchange="onFile(this,'bl-file-name','bl-dz')">
        <label for="bl-file-inp" id="bl-dz" class="dz">
          <div class="di">📊</div>
          <div class="dl">Cliquer ici pour choisir un fichier</div>
          <div class="ds">CSV, XLSX acceptés</div>
          <div class="dn" id="bl-file-name"></div>
        </label>
        <div style="display:flex;gap:8px">
          <button class="btn bp" id="bl-imp-btn" onclick="doImportBL()">⬆ Importer</button>
          <button class="btn bs" onclick="window.open('/api/export/template-budget-lines?year='+document.getElementById('bl-yr-imp').value)">⬇ Template XLSX</button>
        </div>
        <div id="bl-imp-res" style="margin-top:10px"></div>
      </div>
    </div>
  </div>
</div>

<!-- ═══ USERS ═══ -->
<div id="tab-users" class="tab">
  <div style="display:flex;gap:8px;margin-bottom:10px">
    <button class="btn bp" onclick="openUserModal()">＋ Nouvel utilisateur</button>
  </div>
  <div class="tw">
    <table>
      <thead><tr><th>Identifiant</th><th>Nom complet</th><th>Rôle</th><th>Directions assignées</th><th>Email</th><th>Actif</th><th></th></tr></thead>
      <tbody id="u-body"></tbody>
    </table>
  </div>
</div>
</div>

<!-- ═══ MODAL ADD TX ═══ -->
<div class="mbg" id="m-tx">
  <div class="modal">
    <div class="mh"><h3>➕ Nouvelle Transaction</h3><button onclick="closeM('m-tx')" style="background:none;border:none;color:#fff;font-size:20px;cursor:pointer">✕</button></div>
    <div class="mb">
      <div class="fc2">
        <div class="fld"><label>Date *</label><input type="date" id="f-date"></div>
        <div class="fld"><label>Direction *</label><select id="f-dir"><option value="">—</option></select></div>
      </div>
      <div class="fld"><label>Libellé(s) — 1ère ligne = libellé budget, suite = objet/description</label>
        <textarea id="f-libelle" rows="3" placeholder="Ligne 1: Libellé de la ligne budgétaire&#10;Ligne 2+: Objet / description de la dépense"></textarea>
      </div>
      <div class="fld"><label>Intitulé de la commande</label><input type="text" id="f-intitule"></div>
      <div class="fc2">
        <div class="fld"><label>Imputation comptable</label><input type="text" id="f-imp" placeholder="ex: 66410200"></div>
        <div class="fld"><label>Montant (FCFA) *</label><input type="number" id="f-montant" min="0" step="1"></div>
      </div>
      <div class="fc3">
        <div class="fld"><label>Nature</label><select id="f-nature"><option>DEPENSE COURANTE</option><option>DEPENSE DE CAPITAL</option></select></div>
        <div class="fld"><label>Désignation</label><select id="f-desig"><option value="">—</option><option>OM</option><option>MM</option><option>MT</option><option>NC</option><option>BC</option><option>LC</option><option>ADDB</option></select></div>
        <div class="fld"><label>Statut</label><select id="f-status"><option value="validated">Validé</option><option value="pending">En attente</option></select></div>
      </div>
      <div id="m-tx-err"></div>
    </div>
    <div class="mf">
      <button class="btn bs" onclick="closeM('m-tx')">Annuler</button>
      <button class="btn bp" onclick="saveTx()">Enregistrer</button>
    </div>
  </div>
</div>

<!-- ═══ MODAL ADD BL ═══ -->
<div class="mbg" id="m-bl">
  <div class="modal">
    <div class="mh"><h3>➕ Nouvelle Ligne Budgétaire</h3><button onclick="closeM('m-bl')" style="background:none;border:none;color:#fff;font-size:20px;cursor:pointer">✕</button></div>
    <div class="mb">
      <div class="fc2">
        <div class="fld"><label>Année *</label><input type="number" id="f-bl-yr" value="2025"></div>
        <div class="fld"><label>Direction *</label><select id="f-bl-dir"><option value="">—</option></select></div>
      </div>
      <div class="fc2">
        <div class="fld"><label>Imputation *</label><input type="text" id="f-bl-imp" placeholder="ex: 60410000"></div>
        <div class="fld"><label>Budget CP (FCFA) *</label><input type="number" id="f-bl-bud" min="0" step="1"></div>
      </div>
      <div class="fld"><label>Libellé</label><input type="text" id="f-bl-lib"></div>
      <div class="fld"><label>Nature</label><select id="f-bl-nat"><option>DEPENSE COURANTE</option><option>DEPENSE DE CAPITAL</option></select></div>
      <div id="m-bl-err"></div>
    </div>
    <div class="mf">
      <button class="btn bs" onclick="closeM('m-bl')">Annuler</button>
      <button class="btn bp" onclick="saveBL()">Enregistrer</button>
    </div>
  </div>
</div>

<!-- ═══ MODAL ADD USER ═══ -->
<div class="mbg" id="m-user">
  <div class="modal">
    <div class="mh"><h3>👤 Gestion Utilisateur</h3><button onclick="closeM('m-user')" style="background:none;border:none;color:#fff;font-size:20px;cursor:pointer">✕</button></div>
    <div class="mb">
      <div class="fc2">
        <div class="fld"><label>Identifiant *</label><input type="text" id="f-u-id"></div>
        <div class="fld"><label>Mot de passe *</label><input type="password" id="f-u-pw" placeholder="Laisser vide pour ne pas changer"></div>
      </div>
      <div class="fld"><label>Nom complet</label><input type="text" id="f-u-nm"></div>
      <div class="fc2">
        <div class="fld"><label>Rôle</label>
          <select id="f-u-role">
            <option value="agent">Agent (lecture seule)</option>
            <option value="agent_plus">Agent+ (peut importer TX)</option>
            <option value="dcf_sub">Sous-Directeur Budget DCF</option>
            <option value="dcf_dir">Directeur DCF</option>
            <option value="admin">Administrateur (accès total)</option>
          </select>
        </div>
        <div class="fld"><label>Email</label><input type="email" id="f-u-em"></div>
      </div>
      <div class="fld">
        <label>Directions assignées <span style="color:#4a7090;font-weight:400;text-transform:none">(cliquer pour sélectionner — vide = toutes)</span></label>
        <div class="dir-tags" id="f-u-dirs"></div>
      </div>
      <input type="hidden" id="f-u-id-edit" value="">
      <div id="m-u-err"></div>
    </div>
    <div class="mf">
      <button class="btn bs" onclick="closeM('m-user')">Annuler</button>
      <button class="btn bp" onclick="saveUser()">Enregistrer</button>
    </div>
  </div>
</div>

<!-- ═══ PRINT AREA ═══ -->
<div id="print-area" class="print-area"></div>

<script>
// ─── STATE ──────────────────────────────────────────────────────────
const S = {tok:'',role:'',name:'',page:1,userDirs:[],selectedTx:new Set()};
const ALL_DIRS = ["BUM","BUT","BUF","DG","DRH","DICOM","DIRCAB","DCRA","DAMR","DC",
  "DNQ","DAS","DFA","DAJR","DAP","DR","DS","DSPI","DSIR","DOP","DT","DCF","DCRM",
  "DRLM","RRSM","RREM","RROM","RRNOM","RRSOM","RRAM","RRNM","RRENM","DCRF","DRLF",
  "RRSF","RREF","RROF","RRNOF","RRSOF","RRAF","RRNF","RRENF","DCRT","DRLT","RRNOT","RRENT"];

// ─── INIT ────────────────────────────────────────────────────────────
window.addEventListener('DOMContentLoaded', async () => {
  const tok = sessionStorage.getItem('tok');
  if (!tok) { location='/login'; return; }
  S.tok = tok;
  const r = await call('/api/me');
  if (!r) { sessionStorage.clear(); location='/login'; return; }
  const me = await r.json();
  S.role = me.role; S.name = me.name||me.u;
  S.userDirs = me.directions || ALL_DIRS;
  document.getElementById('h-name').textContent = S.name;
  document.getElementById('h-role').textContent = '('+S.role+')';
  // Build year selector
  const sel = document.getElementById('g-yr');
  sel.innerHTML = '';
  (me.years||[new Date().getFullYear()]).forEach(y=>{
    const o=document.createElement('option'); o.value=y; o.textContent=y; sel.appendChild(o);
  });
  // Build direction selector
  buildDirSelector(S.userDirs);
  // Role permissions
  if (S.role==='admin') { document.getElementById('nb-users').style.display=''; document.getElementById('btn-add-bl').style.display=''; }
  if (['admin','dcf_dir','dcf_sub'].includes(S.role)) document.getElementById('btn-add-bl').style.display='';
  const canImport = ['admin','dcf_dir','dcf_sub','agent_plus'].includes(S.role);
  document.getElementById('nb-import').style.display = canImport?'':'none';
  const blBtn = document.getElementById('bl-imp-btn');
  if (blBtn) blBtn.style.display = ['admin','dcf_dir','dcf_sub'].includes(S.role)?'':'none';
  await loadDash();
});

function buildDirSelector(dirs) {
  const sel = document.getElementById('g-dir');
  sel.innerHTML = '<option value="">Toutes</option>';
  dirs.forEach(d => { const o=document.createElement('option'); o.value=d; o.textContent=d; sel.appendChild(o); });
  // Also build form selects
  ['f-dir','f-bl-dir'].forEach(id => {
    const el = document.getElementById(id); if(!el) return;
    el.innerHTML = '<option value="">— Choisir —</option>';
    dirs.forEach(d => { const o=document.createElement('option'); o.value=d; o.textContent=d; el.appendChild(o); });
  });
}

// ─── API CALL ─────────────────────────────────────────────────────────
async function call(path, opts={}) {
  const tok = S.tok||sessionStorage.getItem('tok')||'';
  const headers = {};
  if (tok) headers['Authorization'] = 'Bearer '+tok;
  if (opts.json!==undefined) { headers['Content-Type']='application/json'; opts.body=JSON.stringify(opts.json); delete opts.json; }
  try {
    const r = await fetch(path, {...opts, headers:{...headers,...(opts.headers||{})}, credentials:'include'});
    if (r.status===401) { sessionStorage.clear(); location='/login'; return null; }
    if (!r.ok) { console.error('API',r.status,path); return null; }
    return r;
  } catch(e) { console.error('NET',e.message); return null; }
}

// ─── HELPERS ─────────────────────────────────────────────────────────
const yr    = () => parseInt(document.getElementById('g-yr').value)||2025;
const dir   = () => document.getElementById('g-dir').value||'';
const esc   = s => { const d=document.createElement('div');d.textContent=s||'';return d.innerHTML; };

function onFilters() { reload(); }
function reload() {
  const id = (document.querySelector('.tab.on')||{id:'tab-dashboard'}).id.replace('tab-','');
  show(id);
}
function show(tab) {
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));
  document.querySelectorAll('nav button').forEach(b=>b.classList.remove('on'));
  const t=document.getElementById('tab-'+tab), b=document.getElementById('nb-'+tab);
  if(t) t.classList.add('on'); if(b) b.classList.add('on');
  if(tab==='transactions') loadTxs();
  if(tab==='budget') loadBLs();
  if(tab==='users') loadUsers();
  if(tab==='dashboard') loadDash();
}
function openM(id) { document.getElementById(id).classList.add('open'); }
function closeM(id) { document.getElementById(id).classList.remove('open'); }
async function logout() { await call('/api/logout',{method:'POST'}); sessionStorage.clear(); location='/login'; }

// ─── DASHBOARD ───────────────────────────────────────────────────────
async function loadDash() {
  const r = await call(`/api/dashboard?year=${yr()}&direction=${dir()}`); if(!r) return;
  const d = await r.json();
  document.getElementById('k-bud').textContent   = fmtAmt(d.budget);
  document.getElementById('k-eng').textContent   = fmtAmt(d.engage);
  document.getElementById('k-pend').textContent  = fmtAmt(d.pending);
  document.getElementById('k-dispo').textContent = fmtAmt(d.available);
  document.getElementById('k-bud2').textContent  = 'Budget approuvé';
  document.getElementById('k-eng2').textContent  = (d.pct||0)+'% consommé · '+(d.tx_count||0)+' tx';
  document.getElementById('k-dispo2').textContent= d.available>=0?'✓ Solde positif':'⚠ Dépassement!';
  drawChart(d.monthly||[]);
  const div = document.getElementById('by-dir');
  if (!d.by_dir||!d.by_dir.length) { div.innerHTML='<div style="padding:20px;text-align:center;color:#3a6888">Aucune donnée</div>'; return; }
  div.innerHTML = d.by_dir.sort((a,b)=>b.eng-a.eng).map(r=>{
    const p = r.bud>0?Math.round(r.eng/r.bud*100):0;
    const c = p>100?'#e53935':p>80?'#fb8c00':'#43a047';
    return `<div style="padding:8px 12px;border-bottom:1px solid rgba(0,50,110,.2)">
      <div style="display:flex;justify-content:space-between;margin-bottom:2px">
        <b style="font-size:11px">${r.d}</b><span style="font-size:10px;color:${c};font-weight:700">${p}%</span>
      </div>
      <div style="font-size:10px;color:#4a7090">${fmtAmt(r.eng)} / ${fmtAmt(r.bud)}</div>
      <div class="prg"><div class="prf" style="width:${Math.min(p,100)}%;background:${c}"></div></div>
    </div>`;
  }).join('');
}

function fmtAmt(n) {
  if (n==null) return '0';
  // Space thousands separator
  return Math.round(n).toString().replace(/\B(?=(\d{3})+(?!\d))/g,' ')+' F';
}

function drawChart(data) {
  const c=document.getElementById('ch-monthly'); if(!c) return;
  const W=c.parentElement.clientWidth-16; c.width=W; c.height=160;
  const ctx=c.getContext('2d'); ctx.clearRect(0,0,W,160);
  if(!data.length) return;
  const maxV=Math.max(...data.map(d=>d.val),1), bw=(W-50)/12, chartH=120, top=16;
  ctx.fillStyle='rgba(0,30,80,.3)'; ctx.fillRect(0,0,W,160);
  data.forEach((d,i)=>{
    const x=50+i*bw, h=(d.val/maxV)*chartH, y=top+chartH-h;
    const g=ctx.createLinearGradient(0,y,0,top+chartH);
    g.addColorStop(0,'rgba(0,180,255,.85)'); g.addColorStop(1,'rgba(0,60,180,.25)');
    ctx.fillStyle=g; ctx.fillRect(x+2,y,bw-5,h);
    ctx.fillStyle='rgba(100,160,220,.55)'; ctx.font='9px Segoe UI'; ctx.textAlign='center';
    ctx.fillText(d.label,x+bw/2,top+chartH+13);
    if(d.val>0){
      const lbl=d.val>=1e9?(d.val/1e9).toFixed(1)+'G':d.val>=1e6?(d.val/1e6).toFixed(0)+'M':d.val>=1e3?(d.val/1e3).toFixed(0)+'K':''+d.val;
      ctx.fillStyle='#64b5f6'; ctx.font='8px Segoe UI'; ctx.fillText(lbl,x+bw/2,y-3);
    }
  });
  ctx.fillStyle='#3a6888'; ctx.font='9px Segoe UI'; ctx.textAlign='right';
  const ml=maxV>=1e9?(maxV/1e9).toFixed(1)+'G':maxV>=1e6?(maxV/1e6).toFixed(0)+'M':maxV>=1e3?(maxV/1e3).toFixed(0)+'K':''+maxV;
  ctx.fillText(ml,46,top+8); ctx.fillText('0',46,top+chartH);
}

// ─── TRANSACTIONS ────────────────────────────────────────────────────
let txTimer;
function txSearch(){ clearTimeout(txTimer); txTimer=setTimeout(()=>{S.page=1;loadTxs();},350); }

async function loadTxs(page) {
  if(page) S.page=page;
  const q=document.getElementById('tx-q').value;
  const r=await call(`/api/transactions?year=${yr()}&direction=${dir()}&page=${S.page}&search=${encodeURIComponent(q)}`);
  if(!r) return;
  const d=await r.json();
  document.getElementById('tx-info').textContent=d.total+' transaction(s)';
  const canDel=['admin','dcf_dir','dcf_sub'].includes(S.role);
  const tb=document.getElementById('tx-body');
  if(!d.rows||!d.rows.length){
    tb.innerHTML='<tr><td colspan="10" style="text-align:center;padding:28px;color:#3a6888">Aucune transaction</td></tr>';
  } else {
    tb.innerHTML=d.rows.map(r=>{
      const libLines = (r.libelle||'').split('\n');
      const lib1 = esc(libLines[0]);
      const lib2 = libLines.slice(1).filter(l=>l.trim()).map(l=>`<div style="font-size:10px;color:#7aaccc;margin-top:2px">${esc(l)}</div>`).join('');
      const objet = r.objet?`<div style="font-size:10px;color:#7aaccc;margin-top:2px">${esc(r.objet)}</div>`:'';
      const isSel = S.selectedTx.has(r.id);
      return `<tr class="${isSel?'sel-row':''}" id="tr-${r.id}">
        <td><input type="checkbox" ${isSel?'checked':''} onchange="toggleTx(${r.id},this.checked)"></td>
        <td style="white-space:nowrap">${r.date}</td>
        <td><b>${r.direction}</b></td>
        <td style="max-width:220px"><div style="font-weight:600">${lib1}</div>${lib2}${objet}</td>
        <td style="font-family:monospace;font-size:10px">${r.imputation||'—'}</td>
        <td><span style="font-size:10px;color:${r.nature==='DEPENSE DE CAPITAL'?'#42a5f5':'#78909c'}">${r.nature==='DEPENSE DE CAPITAL'?'CAP':'COUR'}</span></td>
        <td class="amt">${r.montant_fmt}</td>
        <td><span class="${r.statut_budget==='OK'||r.statut_budget==='OUI'?'bok':'bnon'}">${r.statut_budget==='OK'||r.statut_budget==='OUI'?'OUI':'NON'}</span></td>
        <td style="font-size:10px;color:#6a9cc0">${esc(r.created_by_name||r.created_by)}</td>
        <td>${canDel?`<button class="btn bd bsm" onclick="delTx(${r.id})">✕</button>`:''}</td>
      </tr>`;
    }).join('');
  }
  updatePrintBtn();
  const pages=Math.ceil(d.total/d.per_page);
  document.getElementById('tx-pages').innerHTML=pages>1?Array.from({length:pages},(_,i)=>
    `<button class="btn ${i+1===S.page?'bp':'bs'} bsm" onclick="loadTxs(${i+1})">${i+1}</button>`
  ).join(''):'';
}

function toggleTx(id, checked) {
  if(checked) S.selectedTx.add(id); else S.selectedTx.delete(id);
  const row=document.getElementById('tr-'+id);
  if(row) row.className=checked?'sel-row':'';
  updatePrintBtn();
}
function toggleAll(cb) {
  document.querySelectorAll('#tx-body input[type=checkbox]').forEach(c=>{
    const id=parseInt(c.closest('tr').id.replace('tr-',''));
    c.checked=cb.checked; toggleTx(id,cb.checked);
  });
}
function updatePrintBtn() {
  const btn=document.getElementById('btn-print');
  if(btn) btn.style.display=S.selectedTx.size>0?'':'none';
}

function openTxModal() {
  document.getElementById('f-date').value=new Date().toISOString().split('T')[0];
  ['f-intitule','f-libelle','f-imp','f-montant'].forEach(id=>document.getElementById(id).value='');
  document.getElementById('m-tx-err').innerHTML='';
  const d=dir(); if(d) document.getElementById('f-dir').value=d;
  openM('m-tx');
}

async function saveTx() {
  const d=document.getElementById('f-dir').value;
  const libelleRaw=document.getElementById('f-libelle').value.trim();
  const montant=parseFloat(document.getElementById('f-montant').value)||0;
  const err=document.getElementById('m-tx-err');
  if(!d||!libelleRaw||!montant){err.innerHTML='<div class="alrt ar">Direction, libellé et montant obligatoires</div>';return;}
  const lines=libelleRaw.split('\n');
  const libelle=lines[0]; const objet=lines.slice(1).join('\n');
  const r=await call('/api/transactions',{method:'POST',json:{
    date:document.getElementById('f-date').value, direction:d,
    intitule:document.getElementById('f-intitule').value,
    libelle, objet, imputation:document.getElementById('f-imp').value,
    nature:document.getElementById('f-nature').value,
    designation:document.getElementById('f-desig').value,
    status:document.getElementById('f-status').value,
    montant, year:yr()
  }});
  if(!r){err.innerHTML='<div class="alrt ar">Erreur serveur</div>';return;}
  const res=await r.json();
  closeM('m-tx');
  if(res.statut_budget==='NON') setTimeout(()=>alert('⚠️ Budget insuffisant pour cette imputation!'),100);
  loadTxs();
}

async function delTx(id) {
  if(!confirm('Supprimer cette transaction ?')) return;
  const r=await call('/api/transactions/'+id,{method:'DELETE'});
  if(r) loadTxs();
}

// ─── PRINT ───────────────────────────────────────────────────────────
async function printSelected() {
  if(!S.selectedTx.size) return;
  const ids=[...S.selectedTx];
  // Fetch all selected transactions
  const txs=[];
  const r=await call(`/api/transactions?year=${yr()}&direction=${dir()}&page=1`);
  if(!r) return;
  const d=await r.json();
  d.rows.forEach(row=>{ if(ids.includes(row.id)) txs.push(row); });
  if(!txs.length) return;

  // Build print HTML — 2 per A4 page
  let html=`<html><head><title>Fiches CAMTEL</title><style>
    @page{size:A4;margin:15mm}
    body{font-family:'Segoe UI',sans-serif;font-size:11px;color:#000;background:#fff}
    .fiche{border:1.5px solid #1a3a6b;border-radius:6px;padding:14px;margin-bottom:12px;page-break-inside:avoid}
    .fiche-header{display:flex;justify-content:space-between;align-items:center;border-bottom:2px solid #1a3a6b;padding-bottom:8px;margin-bottom:10px}
    .fiche-title{font-size:13px;font-weight:800;color:#1a3a6b}
    .fiche-sub{font-size:10px;color:#555}
    .fiche-grid{display:grid;grid-template-columns:1fr 1fr;gap:6px 16px;margin-bottom:8px}
    .fiche-row{display:flex;flex-direction:column}
    .fiche-label{font-size:9px;font-weight:700;text-transform:uppercase;color:#888;letter-spacing:.04em}
    .fiche-value{font-size:11px;font-weight:600;color:#000;border-bottom:1px dotted #ccc;padding-bottom:2px;min-height:18px}
    .fiche-value.amt{font-family:'Courier New',monospace;font-size:12px}
    .fiche-libelle{margin-bottom:8px}
    .fiche-objet{font-size:10px;color:#444;margin-top:2px}
    .dispo-oui{color:#166534;font-weight:700}.dispo-non{color:#991b1b;font-weight:700}
    .fiche-signatures{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-top:12px;border-top:1px solid #ccc;padding-top:10px}
    .sig-box{text-align:center}
    .sig-line{height:40px;border-bottom:1px solid #888;margin-bottom:4px}
    .sig-label{font-size:9px;color:#555;text-transform:uppercase;letter-spacing:.04em}
    .page-break{page-break-after:always}
  </style></head><body>`;

  txs.forEach((tx,i)=>{
    const libelleLines=(tx.libelle||'').split('\n');
    const lib1=libelleLines[0]||'';
    const lib2=libelleLines.slice(1).filter(l=>l.trim()).join(' — ');
    const objet=tx.objet||lib2||'';
    const dispo=tx.statut_budget==='OK'||tx.statut_budget==='OUI';
    html+=`<div class="fiche">
      <div class="fiche-header">
        <div><div class="fiche-title">CAMTEL — FICHE D'ENGAGEMENT BUDGÉTAIRE</div>
          <div class="fiche-sub">Direction : <b>${tx.direction}</b> &nbsp;|&nbsp; Année : <b>${tx.date?tx.date.substring(0,4):'—'}</b></div></div>
        <div style="text-align:right"><div class="fiche-sub">Date : <b>${tx.date||'—'}</b></div>
          <div class="fiche-sub">Réf : #${tx.id}</div></div>
      </div>
      <div class="fiche-libelle">
        <div class="fiche-label">Libellé de la ligne budgétaire</div>
        <div class="fiche-value">${esc2(lib1)}</div>
        ${objet?`<div class="fiche-objet"><b>Objet :</b> ${esc2(objet)}</div>`:''}
      </div>
      <div class="fiche-grid">
        <div class="fiche-row"><div class="fiche-label">Imputation comptable</div><div class="fiche-value">${tx.imputation||'—'}</div></div>
        <div class="fiche-row"><div class="fiche-label">Nature de la dépense</div><div class="fiche-value">${tx.nature||'—'}</div></div>
        <div class="fiche-row"><div class="fiche-label">Montant (FCFA)</div><div class="fiche-value amt">${tx.montant_fmt||''}</div></div>
        <div class="fiche-row"><div class="fiche-label">Désignation</div><div class="fiche-value">${tx.designation||'—'}</div></div>
        <div class="fiche-row"><div class="fiche-label">Disponible budgétaire</div>
          <div class="fiche-value ${dispo?'dispo-oui':'dispo-non'}">${dispo?'OUI — Budget suffisant':'NON — Dépassement budgétaire'}</div></div>
        <div class="fiche-row"><div class="fiche-label">Initiateur</div><div class="fiche-value">${esc2(tx.created_by_name||tx.created_by||'—')}</div></div>
      </div>
      <div class="fiche-signatures">
        <div class="sig-box"><div class="sig-line"></div><div class="sig-label">Initiateur</div></div>
        <div class="sig-box"><div class="sig-line"></div><div class="sig-label">Avis de la DCF</div></div>
        <div class="sig-box"><div class="sig-line"></div><div class="sig-label">Directeur Général</div></div>
      </div>
    </div>`;
    // Page break every 2 fiches
    if((i+1)%2===0 && i+1<txs.length) html+='<div class="page-break"></div>';
  });
  html+='</body></html>';
  const win=window.open('','_blank');
  win.document.write(html); win.document.close();
  win.onload=()=>{ win.print(); };
}

function esc2(s){ return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

// ─── BUDGET LINES ────────────────────────────────────────────────────
async function loadBLs() {
  const r=await call(`/api/budget-lines?year=${yr()}&direction=${dir()}`); if(!r) return;
  const rows=await r.json();
  document.getElementById('bl-info').textContent=rows.length+' ligne(s)';
  const canDel=S.role==='admin';
  const canEdit=['admin','dcf_dir','dcf_sub'].includes(S.role);
  const tb=document.getElementById('bl-body');
  if(!rows.length){
    tb.innerHTML='<tr><td colspan="10" style="text-align:center;padding:28px;color:#3a6888">Aucune ligne. Importez ou ajoutez manuellement.</td></tr>'; return;
  }
  tb.innerHTML=rows.map(r=>{
    const p=r.pct||0, c=p>100?'#e53935':p>80?'#fb8c00':'#43a047';
    return `<tr>
      <td><b>${r.direction}</b></td>
      <td style="font-family:monospace;font-size:10px">${r.imputation}</td>
      <td style="max-width:200px"><div style="font-weight:600;font-size:11px">${esc(r.libelle)}</div></td>
      <td><span style="font-size:10px;color:${r.nature==='DEPENSE DE CAPITAL'?'#42a5f5':'#78909c'}">${r.nature==='DEPENSE DE CAPITAL'?'CAP':'COUR'}</span></td>
      <td class="amt">${r.budget_cp_fmt}</td>
      <td class="amt">${r.engage_fmt}</td>
      <td class="amt"><span style="color:${r.dispo_ok?'#43a047':'#e53935'}">${r.dispo_ok?'':'-'}${r.dispo_fmt}</span></td>
      <td style="min-width:80px"><span style="color:${c};font-weight:700">${p}%</span>
        <div class="prg" style="width:70px"><div class="prf" style="width:${Math.min(p,100)}%;background:${c}"></div></div></td>
      <td><span class="${r.dispo_ok?'bok':'bnon'}">${r.dispo_ok?'OUI':'NON'}</span></td>
      <td>${canDel?`<button class="btn bd bsm" onclick="delBL(${r.id})">✕</button>`:''}</td>
    </tr>`;
  }).join('');
}

function openBLModal() {
  ['f-bl-imp','f-bl-lib','f-bl-bud'].forEach(id=>document.getElementById(id).value='');
  document.getElementById('f-bl-yr').value=yr();
  document.getElementById('m-bl-err').innerHTML='';
  const d=dir(); if(d) document.getElementById('f-bl-dir').value=d;
  openM('m-bl');
}

async function saveBL() {
  const d=document.getElementById('f-bl-dir').value;
  const imp=document.getElementById('f-bl-imp').value.trim();
  const bud=parseFloat(document.getElementById('f-bl-bud').value)||0;
  const err=document.getElementById('m-bl-err');
  if(!d||!imp||!bud){err.innerHTML='<div class="alrt ar">Direction, imputation et budget obligatoires</div>';return;}
  const r=await call('/api/budget-lines',{method:'POST',json:{
    year:parseInt(document.getElementById('f-bl-yr').value), direction:d, imputation:imp,
    libelle:document.getElementById('f-bl-lib').value, nature:document.getElementById('f-bl-nat').value, budget_cp:bud
  }});
  if(!r){err.innerHTML='<div class="alrt ar">Erreur</div>';return;}
  closeM('m-bl'); loadBLs();
}

async function delBL(id) {
  if(!confirm('Supprimer ?')) return;
  const r=await call('/api/budget-lines/'+id,{method:'DELETE'});
  if(r) loadBLs();
}

// ─── IMPORT ───────────────────────────────────────────────────────────
function onFile(input,fnameId,dzId) {
  const f=input.files[0]; if(!f) return;
  const el=document.getElementById(fnameId); if(el) el.textContent='✓ '+f.name+' ('+(f.size/1024).toFixed(1)+' KB)';
  const dz=document.getElementById(dzId); if(dz){dz.style.borderColor='rgba(0,200,120,.7)';dz.style.background='rgba(0,60,30,.2)';}
}

async function doImportTx() {
  const res=document.getElementById('tx-imp-res');
  const fi=document.getElementById('tx-file-inp'); const f=fi.files[0];
  if(!f){res.innerHTML='<div class="alrt ay">Cliquez sur la zone pour sélectionner un fichier</div>';return;}
  const y=parseInt(document.getElementById('tx-yr-imp').value);
  res.innerHTML='<div class="alrt ab">🔄 Import de <b>'+f.name+'</b>...</div>';
  const fd=new FormData(); fd.append('file',f); fd.append('year',String(y));
  const tok=S.tok||sessionStorage.getItem('tok')||'';
  try{
    const resp=await fetch('/api/import/transactions',{method:'POST',headers:tok?{'Authorization':'Bearer '+tok}:{},credentials:'include',body:fd});
    if(resp.status===401){sessionStorage.clear();location='/login';return;}
    if(resp.status===403){res.innerHTML='<div class="alrt ar">❌ Accès refusé</div>';return;}
    if(!resp.ok){res.innerHTML='<div class="alrt ar">❌ Erreur serveur ('+resp.status+')</div>';return;}
    const d=await resp.json();
    if(d.created>0){
      res.innerHTML=`<div class="alrt ag">✅ ${d.created} transaction(s) importée(s) pour ${y}</div>`;
      fi.value=''; document.getElementById('tx-file-name').textContent='';
      const dz=document.getElementById('tx-dz'); dz.style.borderColor=''; dz.style.background='';
      const sel=document.getElementById('g-yr');
      if(!Array.from(sel.options).find(o=>parseInt(o.value)===y)){const opt=document.createElement('option');opt.value=y;opt.textContent=y;sel.appendChild(opt);}
      sel.value=y; await loadDash();
    } else {
      res.innerHTML=`<div class="alrt ay">⚠ Aucune transaction importée. Vérifiez le fichier.</div>`;
    }
  }catch(e){res.innerHTML='<div class="alrt ar">❌ Erreur: '+e.message+'</div>';}
}

async function doImportBL() {
  const res=document.getElementById('bl-imp-res');
  const fi=document.getElementById('bl-file-inp'); const f=fi.files[0];
  if(!f){res.innerHTML='<div class="alrt ay">Cliquez sur la zone pour sélectionner un fichier</div>';return;}
  const y=parseInt(document.getElementById('bl-yr-imp').value)||0;
  res.innerHTML='<div class="alrt ab">🔄 Import de <b>'+f.name+'</b>...</div>';
  const fd=new FormData(); fd.append('file',f); fd.append('year',String(y));
  const tok=S.tok||sessionStorage.getItem('tok')||'';
  try{
    const resp=await fetch('/api/import/budget-lines',{method:'POST',headers:tok?{'Authorization':'Bearer '+tok}:{},credentials:'include',body:fd});
    if(resp.status===401){sessionStorage.clear();location='/login';return;}
    if(resp.status===403){res.innerHTML='<div class="alrt ar">❌ Accès refusé (admin/dcf requis)</div>';return;}
    if(!resp.ok){res.innerHTML='<div class="alrt ar">❌ Erreur ('+resp.status+')</div>';return;}
    const d=await resp.json();
    if(d.created>0||d.updated>0){
      res.innerHTML=`<div class="alrt ag">✅ ${d.created} créée(s), ${d.updated} mise(s) à jour</div>`;
      fi.value=''; document.getElementById('bl-file-name').textContent='';
      const dz=document.getElementById('bl-dz'); dz.style.borderColor=''; dz.style.background='';
      await loadBLs();
    } else {
      res.innerHTML=`<div class="alrt ay">⚠ Aucune ligne importée. Vérifiez le fichier.</div>`;
    }
  }catch(e){res.innerHTML='<div class="alrt ar">❌ Erreur: '+e.message+'</div>';}
}

// ─── USERS ────────────────────────────────────────────────────────────
async function loadUsers() {
  const r=await call('/api/users'); if(!r) return;
  const users=await r.json();
  document.getElementById('u-body').innerHTML=users.map(u=>`<tr>
    <td><b>${esc(u.username)}</b></td>
    <td>${esc(u.full_name)}</td>
    <td><span class="bok" style="font-size:10px">${u.role}</span></td>
    <td style="max-width:200px;font-size:10px;color:#7aaccc">${(u.directions||[]).length===0?'<i>Toutes</i>':(u.directions||[]).join(', ')}</td>
    <td>${esc(u.email)||'—'}</td>
    <td>${u.is_active?'✓':'✗'}</td>
    <td style="display:flex;gap:4px">
      <button class="btn bs bsm" onclick="editUser(${JSON.stringify(u).replace(/"/g,'&quot;')})">✏</button>
      <button class="btn bd bsm" onclick="delUser(${u.id})">✕</button>
    </td>
  </tr>`).join('');
}

function buildDirTags(selected=[]) {
  const container=document.getElementById('f-u-dirs');
  container.innerHTML='';
  ALL_DIRS.forEach(d=>{
    const tag=document.createElement('span');
    tag.className='dir-tag'+(selected.includes(d)?' sel':'');
    tag.textContent=d;
    tag.onclick=()=>{ tag.classList.toggle('sel'); };
    container.appendChild(tag);
  });
}

function getSelectedDirs() {
  return [...document.querySelectorAll('#f-u-dirs .dir-tag.sel')].map(t=>t.textContent);
}

function openUserModal() {
  document.getElementById('f-u-id').value=''; document.getElementById('f-u-id').disabled=false;
  document.getElementById('f-u-pw').value=''; document.getElementById('f-u-nm').value='';
  document.getElementById('f-u-em').value=''; document.getElementById('f-u-id-edit').value='';
  document.getElementById('m-u-err').innerHTML='';
  document.getElementById('f-u-role').value='agent';
  buildDirTags([]);
  openM('m-user');
}

function editUser(u) {
  document.getElementById('f-u-id').value=u.username; document.getElementById('f-u-id').disabled=true;
  document.getElementById('f-u-pw').value=''; document.getElementById('f-u-nm').value=u.full_name||'';
  document.getElementById('f-u-em').value=u.email||''; document.getElementById('f-u-id-edit').value=u.id;
  document.getElementById('f-u-role').value=u.role||'agent';
  document.getElementById('m-u-err').innerHTML='';
  buildDirTags(u.directions||[]);
  openM('m-user');
}

async function saveUser() {
  const username=document.getElementById('f-u-id').value.trim();
  const pw=document.getElementById('f-u-pw').value;
  const editId=document.getElementById('f-u-id-edit').value;
  const err=document.getElementById('m-u-err');
  if(!username){err.innerHTML='<div class="alrt ar">Identifiant requis</div>';return;}
  if(!editId && !pw){err.innerHTML='<div class="alrt ar">Mot de passe requis pour un nouvel utilisateur</div>';return;}
  const dirs=getSelectedDirs();
  const payload={username,full_name:document.getElementById('f-u-nm').value,
    role:document.getElementById('f-u-role').value,
    email:document.getElementById('f-u-em').value,directions:dirs};
  if(pw) payload.password=pw;
  let r;
  if(editId){
    r=await call('/api/users/'+editId,{method:'PUT',json:payload});
  } else {
    if(!pw){err.innerHTML='<div class="alrt ar">Mot de passe requis</div>';return;}
    r=await call('/api/users',{method:'POST',json:payload});
  }
  if(!r){err.innerHTML='<div class="alrt ar">Erreur (identifiant déjà existant?)</div>';return;}
  closeM('m-user'); loadUsers();
}

async function delUser(id) {
  if(!confirm('Supprimer cet utilisateur ?')) return;
  const r=await call('/api/users/'+id,{method:'DELETE'});
  if(r) loadUsers();
}
</script>
</body>
</html>"""
